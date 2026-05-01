import io
import re
import unicodedata
from typing import Any, Dict, List, Tuple

import pandas as pd
import streamlit as st
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Wholesale Lead Quality Checker", page_icon="✅", layout="wide")

HEADER_YELLOW = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
CELL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
CELL_BLUE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
CELL_AMBER = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CELL_RED = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

DASH_CHARS = "\u2010\u2011\u2012\u2013\u2014\u2212"

def norm_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value)
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\u00A0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text.casefold()

def norm_key(value: Any) -> str:
    text = norm_text(value)
    text = text.replace("&", " and ")
    for ch in DASH_CHARS:
        text = text.replace(ch, " ")
    text = re.sub(r"[\\/]", " ", text)
    text = re.sub(r"[^a-z0-9+#.\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def clean_header(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\u00A0", " ")
    text = text.strip().lower()
    text = text.replace("_", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

PLACEHOLDER_VALUES = {
    "", "picklist", "leave blank", "blank", "integer", "text", "date",
    "dd mm yyyy", "mm dd yyyy", "https www", "http www", "do not map",
}

def is_placeholder(value: Any) -> bool:
    text = norm_key(value)
    if text in PLACEHOLDER_VALUES:
        return True
    raw = norm_text(value)
    return (
        raw.startswith("all accepted")
        or raw.startswith("target the below")
        or raw.startswith("no proof")
        or raw.startswith("no toll")
        or raw.startswith("no fee")
    )

def looks_like_template_row(row: pd.Series) -> bool:
    values = [str(v).strip() for v in row.tolist() if str(v).strip()]
    if not values:
        return True
    joined = " | ".join(values[:40]).casefold()
    markers = [
        "picklist", "leave blank", "integer", "text", "dd/mm/yyyy",
        "mm/dd/yyyy", "https://", "no toll", "fee phone",
        "target the below", "all accepted", "do not map",
    ]
    hits = sum(1 for marker in markers if marker in joined)
    if hits >= 2:
        return True
    placeholder_count = sum(1 for value in values if is_placeholder(value))
    return bool(values and placeholder_count / len(values) >= 0.6)

SYNONYM_GROUPS = [
    ["us", "usa", "u s", "u s a", "united states", "united states of america"],
    ["uk", "u k", "gb", "gbr", "great britain", "united kingdom", "england"],
    ["it", "information technology", "technology"],
    ["it operations", "information technology operations", "it ops", "technology operations"],
    ["hr", "human resources"],
    ["vp", "vice president"],
    ["svp", "senior vice president"],
    ["evp", "executive vice president"],
    ["c level", "c-level", "c suite", "c-suite", "chief"],
    ["biz dev", "business development"],
    ["ops", "operations"],
]

def canonical_key(value: Any) -> str:
    key = norm_key(value)
    for group in SYNONYM_GROUPS:
        keys = [norm_key(item) for item in group]
        if key in keys:
            return keys[0]
    return key

def expanded_keys(value: Any) -> List[str]:
    key = norm_key(value)
    canonical = canonical_key(value)
    keys = []
    for item in [key, canonical]:
        if item and item not in keys:
            keys.append(item)
    for group in SYNONYM_GROUPS:
        group_keys = [norm_key(item) for item in group]
        if key in group_keys or canonical in group_keys:
            for item in group_keys:
                if item and item not in keys:
                    keys.append(item)
    return keys

FIELD_ALIASES: Dict[str, List[str]] = {
    "company": ["company", "company name", "account", "account name", "organisation", "organization"],
    "email": ["email", "email address", "work email"],
    "country": ["country", "lead country", "company country"],
    "region": ["region", "state", "province", "c state"],
    "industry": ["industry", "companyindustry", "company industry", "main industry", "indcode1", "c industry"],
    "sub_industry": ["sub industry", "subindustry", "indcode2"],
    "function": ["function", "department", "departments", "job function", "job area", "ocpcode1", "ocpcode2"],
    "job_level": ["position", "job level", "job_level", "seniority", "level", "ocpcode3"],
    "job_role": ["job role", "role", "job_role", "job_role__c"],
    "company_size": ["companysize", "company size", "number of employees", "number_of_employees", "employees", "orgemp"],
    "phone": ["phone", "telephone", "mobile", "cell", "phonemain", "phone main"],
    "website": ["website", "domain", "url", "companyurl", "company url", "web"],
    "job_title": ["job title", "job_title", "jobtitle", "title", "jobtitletext", "job title text"],
}

def detect_columns(df: pd.DataFrame) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for field, aliases in FIELD_ALIASES.items():
        best_col = ""
        best_score = 0
        for col in df.columns:
            header = clean_header(col)
            if field == "country" and "county" in header.split():
                continue
            for alias in aliases:
                alias_clean = clean_header(alias)
                if header == alias_clean:
                    score = 100
                elif alias_clean in header:
                    score = 85
                else:
                    score = fuzz.token_sort_ratio(header, alias_clean)
                if score > best_score:
                    best_score = score
                    best_col = str(col)
        if best_col and best_score >= 72:
            result[field] = best_col
    return result

def field_for_picklist_column(column_name: Any) -> str:
    header = clean_header(column_name)
    if not header or "county" in header.split():
        return ""
    best_field = ""
    best_score = 0
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            alias_clean = clean_header(alias)
            if header == alias_clean:
                score = 100
            elif alias_clean in header:
                score = 85
            else:
                score = fuzz.token_sort_ratio(header, alias_clean)
            if score > best_score:
                best_score = score
                best_field = field
    return best_field if best_score >= 72 else ""

def is_value_or_code_column(column_name: Any) -> bool:
    header = clean_header(column_name)
    return header == "value" or header.startswith("value ") or header == "code" or header.endswith(" code")

def add_allowed(allowed: Dict[str, Dict[str, str]], field: str, value: Any) -> None:
    if not field or is_placeholder(value):
        return
    text = str(value).strip()
    if not text or len(text) > 120:
        return
    allowed.setdefault(field, {})
    for key in expanded_keys(text):
        allowed[field].setdefault(key, text)

def extract_picklist_rules(df_pick: pd.DataFrame) -> Dict[str, Any]:
    allowed: Dict[str, Dict[str, str]] = {}
    mapping_pairs = []
    columns = list(df_pick.columns)
    for col in columns:
        field = field_for_picklist_column(col)
        if not field:
            continue
        for value in df_pick[col].tolist():
            add_allowed(allowed, field, value)
    for index, col in enumerate(columns):
        field = field_for_picklist_column(col)
        if not field:
            continue
        nearby = columns[index + 1:index + 4]
        for partner in nearby:
            if is_value_or_code_column(partner):
                mapping_pairs.append((str(col), str(partner), field))
                for label, code in zip(df_pick[col].tolist(), df_pick[partner].tolist()):
                    add_allowed(allowed, field, label)
                    add_allowed(allowed, field, code)
                break
    return {
        "allowed": allowed,
        "allowed_counts": {field: len(values) for field, values in allowed.items()},
        "mapping_pairs": mapping_pairs,
        "samples": {field: list(values.values())[:10] for field, values in allowed.items()},
    }

def match_value(value: Any, allowed_map: Dict[str, str], allow_fuzzy: bool = True) -> Tuple[str, str, int]:
    raw = str(value or "").strip()
    if not raw:
        return "Review", "blank value", 0
    if not allowed_map:
        return "Rule Missing", "no picklist values detected", 0
    for key in expanded_keys(raw):
        if key in allowed_map:
            return "Match", f"matched to {allowed_map[key]}", 100
    if allow_fuzzy:
        match = process.extractOne(norm_key(raw), list(allowed_map.keys()), scorer=fuzz.token_sort_ratio)
        if match:
            matched_key = match[0]
            score = int(match[1])
            if score >= 90:
                return "Match", f"strong fuzzy match to {allowed_map[matched_key]} ({score})", score
            if score >= 75:
                return "Review", f"possible fuzzy match to {allowed_map[matched_key]} ({score})", score
    return "No Match", "not found in picklist", 0

COMPANY_SUFFIXES = {
    "ltd", "limited", "co", "company", "corp", "corporation", "inc", "incorporated",
    "plc", "llc", "sa", "ag", "nv", "se", "bv", "oy", "ab", "aps", "as", "sarl",
    "sas", "spa", "gmbh", "pte", "pty", "sdn", "bhd", "holdings", "holding", "group",
}

def email_domain(email: Any) -> str:
    text = str(email or "").strip()
    if "@" not in text:
        return ""
    return text.split("@", 1)[1].strip().lower()

def clean_domain(domain: Any) -> str:
    text = str(domain or "").strip().lower()
    text = re.sub(r"^https?://", "", text)
    text = re.sub(r"/.*$", "", text)
    text = text.replace("www.", "")
    return text

def domain_base(domain: Any) -> str:
    text = clean_domain(domain)
    if not text:
        return ""
    return re.sub(r"[^a-z0-9]", "", text.split(".")[0])

def company_tokens(company: Any) -> List[str]:
    text = str(company or "")
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"[^A-Za-z0-9\s]", " ", text)
    tokens = [token.lower() for token in text.split() if token.strip()]
    tokens = [token for token in tokens if token not in COMPANY_SUFFIXES]
    tokens = [token for token in tokens if token not in {"of", "and", "the", "for", "to", "a"}]
    return tokens

def compare_company_domain(company: Any, email: Any, website: Any) -> Tuple[str, str, int]:
    company_text = str(company or "").strip()
    domain = email_domain(email) or clean_domain(website)
    if not company_text or not domain:
        return "Review", "missing company or domain", 0
    base = domain_base(domain)
    if not base:
        return "Review", "invalid domain", 0
    tokens = company_tokens(company_text)
    joined = "".join(tokens)
    if joined and joined in base:
        return "Match", "company tokens contained in domain", 95
    score = max(
        fuzz.token_sort_ratio(" ".join(tokens), base),
        fuzz.partial_ratio(" ".join(tokens), base),
    )
    score = int(score)
    if score >= 85:
        return "Match", f"strong fuzzy company/domain match ({score})", score
    if score >= 70:
        return "Review", f"weak fuzzy company/domain match ({score})", score
    return "No Match", f"low company/domain similarity ({score})", score

TITLE_TERMS = [
    "chief", "ceo", "cfo", "cio", "cto", "coo", "ciso",
    "president", "vice president", "vp", "svp", "evp",
    "director", "head", "manager",
    "technology", "information technology", "operations",
    "digital", "data", "security", "infrastructure", "systems",
]

def title_relevance(title: Any, function: Any, level: Any) -> Tuple[str, str, int]:
    combined = f" {norm_key(title)} {norm_key(function)} {norm_key(level)} "
    if not combined.strip():
        return "Review", "missing title/function/level", 0
    hits = []
    score = 0
    for term in TITLE_TERMS:
        if f" {norm_key(term)} " in combined:
            hits.append(term)
            if term in {"chief", "ceo", "cfo", "cio", "cto", "coo", "ciso"}:
                score = max(score, 100)
            elif term in {"president", "vice president", "vp", "svp", "evp"}:
                score = max(score, 90)
            elif term in {"director", "head"}:
                score = max(score, 75)
            elif term == "manager":
                score = max(score, 60)
            else:
                score = max(score, 65)
    if score >= 85:
        return "Match", "strong target signal: " + ", ".join(sorted(set(hits))), score
    if score >= 60:
        return "Review", "possible target signal: " + ", ".join(sorted(set(hits))), score
    return "No Match", "no clear target title signal", score

WEIGHTS = {
    "country": 15,
    "industry": 15,
    "company_size": 10,
    "function": 15,
    "job_level": 15,
    "title": 15,
    "domain": 10,
    "phone": 5,
}

def points_for_status(status: str, weight: int) -> float:
    status_clean = norm_key(status)
    if status_clean == "match":
        return float(weight)
    if status_clean in {"review", "rule missing", "column missing"}:
        return weight * 0.5
    return 0.0

def overall_status(score: float) -> str:
    if score >= 85:
        return "PASS"
    if score >= 60:
        return "REVIEW"
    return "FAIL"

def fill_for_value(value: Any) -> PatternFill:
    text = norm_key(value)
    if text in {"pass", "match", "yes"}:
        return CELL_GREEN
    if text in {"review", "rule missing", "column missing"}:
        return CELL_AMBER
    if text in {"fail", "no match", "no"}:
        return CELL_RED
    if not text:
        return CELL_GREEN
    return CELL_BLUE

QA_COLUMNS = [
    "QA_Country_Status", "QA_Country_Reason", "QA_Country_Score",
    "QA_Industry_Status", "QA_Industry_Reason", "QA_Industry_Score",
    "QA_Company_Size_Status", "QA_Company_Size_Reason", "QA_Company_Size_Score",
    "QA_Function_Status", "QA_Function_Reason", "QA_Function_Score",
    "QA_Job_Level_Status", "QA_Job_Level_Reason", "QA_Job_Level_Score",
    "QA_Title_Relevance_Status", "QA_Title_Relevance_Reason", "QA_Title_Relevance_Score",
    "QA_Domain_Status", "QA_Domain_Reason", "QA_Domain_Score",
    "QA_Score", "QA_Overall_Status", "QA_Issues", "QA_Debug_Notes",
]

def write_results_to_workbook(
    master_bytes: bytes,
    sheet_name: str,
    results: pd.DataFrame,
    qa_columns: List[str],
    apply_colours: bool,
) -> bytes:
    workbook = load_workbook(io.BytesIO(master_bytes))
    worksheet = workbook[sheet_name]
    worksheet.title = "Results"
    start_col = worksheet.max_column + 1
    for index, column_name in enumerate(qa_columns):
        col_number = start_col + index
        worksheet.cell(row=1, column=col_number).value = column_name
        if apply_colours:
            worksheet.cell(row=1, column=col_number).fill = HEADER_YELLOW
    column_positions = {column_name: start_col + index for index, column_name in enumerate(qa_columns)}
    for _, result_row in results.iterrows():
        excel_row = int(result_row["_excel_row"])
        for column_name in qa_columns:
            col_number = column_positions[column_name]
            value = result_row.get(column_name, "")
            worksheet.cell(row=excel_row, column=col_number).value = value
            if apply_colours:
                if column_name.endswith("_Status") or column_name == "QA_Overall_Status":
                    worksheet.cell(row=excel_row, column=col_number).fill = fill_for_value(value)
                elif column_name == "QA_Score":
                    try:
                        numeric_score = float(value)
                        if numeric_score >= 85:
                            worksheet.cell(row=excel_row, column=col_number).fill = CELL_GREEN
                        elif numeric_score >= 60:
                            worksheet.cell(row=excel_row, column=col_number).fill = CELL_AMBER
                        else:
                            worksheet.cell(row=excel_row, column=col_number).fill = CELL_RED
                    except Exception:
                        worksheet.cell(row=excel_row, column=col_number).fill = CELL_BLUE
                elif column_name == "QA_Issues":
                    worksheet.cell(row=excel_row, column=col_number).fill = CELL_GREEN if not value else CELL_AMBER
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()

def process_file(
    master_bytes: bytes,
    picklist_bytes: bytes,
    master_sheet: str,
    picklist_sheet: str,
    apply_colours: bool,
) -> Tuple[bytes, Dict[str, Any]]:
    master_df = pd.read_excel(io.BytesIO(master_bytes), sheet_name=master_sheet, dtype=str, keep_default_na=False)
    picklist_df = pd.read_excel(io.BytesIO(picklist_bytes), sheet_name=picklist_sheet, dtype=str, keep_default_na=False)
    colmap = detect_columns(master_df)
    rules = extract_picklist_rules(picklist_df)
    results = []
    skipped_rows = []

    def get_value(row: pd.Series, field: str) -> str:
        column = colmap.get(field, "")
        if column and column in master_df.columns:
            return str(row.get(column, "") or "").strip()
        return ""

    for index, row in master_df.iterrows():
        excel_row = index + 2
        if looks_like_template_row(row):
            skipped_rows.append(excel_row)
            continue
        result: Dict[str, Any] = {"_excel_row": excel_row}
        issues = []
        total_score = 0.0
        field_checks = [
            ("country", "QA_Country", WEIGHTS["country"], False),
            ("industry", "QA_Industry", WEIGHTS["industry"], True),
            ("company_size", "QA_Company_Size", WEIGHTS["company_size"], True),
            ("function", "QA_Function", WEIGHTS["function"], True),
            ("job_level", "QA_Job_Level", WEIGHTS["job_level"], True),
        ]
        for field, prefix, weight, allow_fuzzy in field_checks:
            if field not in colmap:
                status, reason, score = "Column Missing", "master column not detected", 0
            else:
                allowed_map = rules["allowed"].get(field, {})
                status, reason, score = match_value(get_value(row, field), allowed_map, allow_fuzzy)
            result[f"{prefix}_Status"] = status
            result[f"{prefix}_Reason"] = reason
            result[f"{prefix}_Score"] = score
            total_score += points_for_status(status, weight)
            if status != "Match":
                issues.append(field)

        title_status, title_reason, title_score = title_relevance(
            get_value(row, "job_title"),
            get_value(row, "function"),
            get_value(row, "job_level"),
        )
        result["QA_Title_Relevance_Status"] = title_status
        result["QA_Title_Relevance_Reason"] = title_reason
        result["QA_Title_Relevance_Score"] = title_score
        total_score += points_for_status(title_status, WEIGHTS["title"])
        if title_status != "Match":
            issues.append("title")

        domain_status, domain_reason, domain_score = compare_company_domain(
            get_value(row, "company"),
            get_value(row, "email"),
            get_value(row, "website"),
        )
        result["QA_Domain_Status"] = domain_status
        result["QA_Domain_Reason"] = domain_reason
        result["QA_Domain_Score"] = domain_score
        total_score += points_for_status(domain_status, WEIGHTS["domain"])
        if domain_status != "Match":
            issues.append("domain")

        result["QA_Score"] = round(total_score, 1)
        result["QA_Overall_Status"] = overall_status(total_score)
        result["QA_Issues"] = "; ".join(sorted(set(issues)))
        result["QA_Debug_Notes"] = f"Detected columns: {colmap}"
        results.append(result)

    results_df = pd.DataFrame(results)
    if results_df.empty:
        results_df = pd.DataFrame(columns=["_excel_row"] + QA_COLUMNS)

    output_bytes = write_results_to_workbook(
        master_bytes=master_bytes,
        sheet_name=master_sheet,
        results=results_df,
        qa_columns=QA_COLUMNS,
        apply_colours=apply_colours,
    )
    debug = {
        "detected_columns": colmap,
        "allowed_counts": rules["allowed_counts"],
        "mapping_pairs": rules["mapping_pairs"],
        "picklist_samples": rules["samples"],
        "skipped_template_rows": skipped_rows,
        "processed_rows": len(results_df),
    }
    return output_bytes, debug

st.title("Wholesale Lead Quality Checker")
st.caption(
    "Standalone wholesale QA tool with template-row skipping, fixed picklist value handling, "
    "synonym/fuzzy matching, scoring, colour coding, and Excel output preservation."
)

master_file = st.file_uploader("Upload Wholesale Master (.xlsx)", type=["xlsx"])
picklist_file = st.file_uploader("Upload Wholesale Picklist (.xlsx)", type=["xlsx"])

master_sheet = None
picklist_sheet = None
master_bytes = None
picklist_bytes = None

if master_file is not None:
    master_bytes = master_file.read()
    master_workbook = load_workbook(io.BytesIO(master_bytes), read_only=True)
    master_sheet = st.selectbox("Master sheet to process", master_workbook.sheetnames)

if picklist_file is not None:
    picklist_bytes = picklist_file.read()
    picklist_workbook = load_workbook(io.BytesIO(picklist_bytes), read_only=True)
    picklist_sheet = st.selectbox("Picklist sheet to use", picklist_workbook.sheetnames)

apply_colours = st.toggle("Colour-code Excel results", value=True)
show_debug = st.toggle("Show debug panel", value=True)

if "output_bytes" not in st.session_state:
    st.session_state.output_bytes = None

if "debug_info" not in st.session_state:
    st.session_state.debug_info = None

can_run = master_bytes is not None and picklist_bytes is not None and master_sheet and picklist_sheet

if st.button("Run Wholesale QA", type="primary", use_container_width=True, disabled=not can_run):
    try:
        with st.spinner("Processing..."):
            output_bytes, debug_info = process_file(
                master_bytes=master_bytes,
                picklist_bytes=picklist_bytes,
                master_sheet=master_sheet,
                picklist_sheet=picklist_sheet,
                apply_colours=apply_colours,
            )
        st.session_state.output_bytes = output_bytes
        st.session_state.debug_info = debug_info
        st.success("Processing complete.")
    except Exception as error:
        st.session_state.output_bytes = None
        st.session_state.debug_info = None
        st.error(f"Error: {error}")

if st.session_state.debug_info and show_debug:
    with st.expander("Debug information", expanded=True):
        st.json(st.session_state.debug_info)

if st.session_state.output_bytes:
    st.download_button(
        label="Download Wholesale QA Results",
        data=st.session_state.output_bytes,
        file_name="wholesale_qa_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
